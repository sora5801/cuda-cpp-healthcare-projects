#!/usr/bin/env python3
# =============================================================================
# tools/catalog.py  --  Build the machine-readable catalog (catalog.json)
# -----------------------------------------------------------------------------
# WHAT THIS IS
#   The repository's two human-facing "source of truth" files are:
#     * CUDA_CPP_Healthcare_Projects.xlsx        (the sortable catalog; 1 row/project)
#     * CUDA_CPP_Healthcare_Projects_DeepDive.md (the same 301 projects in long form)
#   Office/Markdown files are awkward for tools to parse repeatedly, so this
#   script flattens BOTH into a single, stable, machine-readable `catalog.json`
#   that every other tool (scaffold.py, status.py, verify_project.py) consumes.
#   Think of catalog.json as the compiled "interface" between the catalog and
#   the automation -- generate it once, and nothing else has to touch xlsx/md.
#
# WHY BOTH SOURCES
#   * The .xlsx gives us clean, column-structured fields (difficulty, datasets,
#     ...), which is what we key the automation on.
#   * The .md gives us the richer long-form prose block per project, which we
#     attach verbatim as `deepdive_md` so THEORY.md authors have the full text
#     at hand without re-opening the markdown.
#
# OUTPUT RECORD (one per project), e.g.:
#   {
#     "section": 1, "domain": "Drug Discovery & Molecular Design",
#     "domain_slug": "01-drug-discovery",
#     "id": "1.1", "id_padded": "1.01",
#     "project": "Molecular Dynamics Engine",
#     "project_slug": "molecular-dynamics-engine",
#     "difficulty": "Beginner", "difficulty_emoji": "U+1F7E2",
#     "maturity": "Established",
#     "deep_dive": "...", "algorithms": "...", "datasets": "...",
#     "repos": "...", "cuda_gpu": "...",
#     "deepdive_md": "### 1.1 Molecular Dynamics Engine ...",
#     "folder_path": "projects/01-drug-discovery/1.01-molecular-dynamics-engine"
#   }
#
# DEPENDENCIES
#   openpyxl is preferred (the contract names it). If it is missing, we fall
#   back to a tiny dependency-free .xlsx reader built on the standard library
#   (zipfile + xml) so the catalog can ALWAYS be regenerated on a clean machine.
#
# USAGE
#   python tools/catalog.py            # writes <repo>/catalog.json
#   python tools/catalog.py --check    # parse + validate only, write nothing
# =============================================================================

import argparse
import json
import re
import sys
import unicodedata
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

# --- Repo geometry -----------------------------------------------------------
# This file lives at <repo>/tools/catalog.py, so the repo root is two levels up.
ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "CUDA_CPP_Healthcare_Projects.xlsx"
DEEPDIVE = ROOT / "CUDA_CPP_Healthcare_Projects_DeepDive.md"
OUT = ROOT / "catalog.json"

# Expected totals, used as a tripwire so a silently-truncated parse fails loudly.
EXPECTED_PROJECTS = 301
EXPECTED_SECTIONS = 14

# --- Canonical domain folder slugs (CLAUDE.md section 3) ---------------------
# These are FIXED per section number -- they are NOT derived from the domain
# name -- so that folders sort 01..14 regardless of wording. Keep in lockstep
# with CLAUDE.md if the domain list ever changes.
DOMAIN_SLUGS = {
    1: "01-drug-discovery",
    2: "02-structural-biology",
    3: "03-genomics",
    4: "04-medical-imaging",
    5: "05-radiation-therapy-medphys",
    6: "06-physiology-systems-biology",
    7: "07-medical-ai",
    8: "08-neuroscience-bci",
    9: "09-epidemiology-public-health",
    10: "10-biomechanics-devices",
    11: "11-biotech-synthbio",
    12: "12-omics-data-processing",
    13: "13-pharmacology-quant",
    14: "14-emerging-frontiers",
}

# Difficulty -> the legend emoji used across READMEs (stored as text, not glyph,
# to keep this file pure-ASCII and portable; scaffold renders the glyph).
DIFF_EMOJI = {"Beginner": "U+1F7E2", "Intermediate": "U+1F7E1", "Advanced": "U+1F534"}


# --- Helpers -----------------------------------------------------------------
def slugify(text: str) -> str:
    """Turn a project name into an ASCII, lowercase, hyphen-joined slug.

    Rule (CLAUDE.md section 3): lowercase, ASCII only, every run of
    non-alphanumeric characters (spaces, '/', punctuation) collapses to a
    single '-', with no leading/trailing '-'. Deterministic so scaffold.py
    always produces identical folder names.

        "Quantum Chemistry / DFT on GPU" -> "quantum-chemistry-dft-on-gpu"
    """
    # NFKD + ascii-encode strips accents (e.g. "Nose-Hoover" stays clean).
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)  # any non-alnum run -> single dash
    return text.strip("-")


def pad_id(id_str: str) -> str:
    """Zero-pad the MINOR part of an ID to two digits so folders sort right.

        "1.1"  -> "1.01"      "3.10" -> "3.10"      "14.5" -> "14.05"
    The major part is left as-is (sections are listed numerically elsewhere).
    """
    major, minor = id_str.split(".")
    return f"{int(major)}.{int(minor):02d}"


# --- .xlsx reading: preferred (openpyxl) + dependency-free fallback ----------
def read_xlsx_openpyxl(path: Path):
    """Read the 'Projects' sheet via openpyxl. Returns (headers, rows)."""
    from openpyxl import load_workbook

    # read_only streams rows (low memory); data_only returns cached values, not
    # formulas (irrelevant for the Projects sheet, which is plain text anyway).
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Projects"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    body = [list(r) for r in rows[1:] if any(c is not None for c in r)]
    return headers, body


def read_xlsx_stdlib(path: Path):
    """Dependency-free .xlsx reader (zipfile + xml) used if openpyxl is absent.

    An .xlsx is a zip of XML parts. We only need:
      * xl/sharedStrings.xml  -- the string pool cells point into
      * xl/worksheets/sheetN.xml for the 'Projects' sheet
    We resolve the sheet name -> file via xl/workbook.xml + its rels. This is a
    deliberately small, well-commented reimplementation so a learner can see
    exactly what openpyxl does under the hood.
    """
    ns = {
        "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    with zipfile.ZipFile(path) as z:
        # 1) shared string table: cells of type "s" carry an index into this.
        shared = []
        if "xl/sharedStrings.xml" in z.namelist():
            sst = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for si in sst.findall("m:si", ns):
                # A string item is either a single <t> or several <r><t> runs.
                shared.append("".join(t.text or "" for t in si.iter("{%s}t" % ns["m"])))

        # 2) map sheet name -> internal part path via workbook.xml + rels.
        wb_xml = ET.fromstring(z.read("xl/workbook.xml"))
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {
            rel.get("Id"): rel.get("Target") for rel in rels.findall("rel:Relationship", ns)
        }
        name_to_part = {}
        for sh in wb_xml.find("m:sheets", ns).findall("m:sheet", ns):
            rid = sh.get("{%s}id" % ns["r"])
            target = rid_to_target[rid]
            part = target if target.startswith("xl/") else "xl/" + target.lstrip("/")
            name_to_part[sh.get("name")] = part

        # 3) parse the 'Projects' worksheet into a dense list-of-lists grid.
        sheet = ET.fromstring(z.read(name_to_part["Projects"]))

    def col_index(cell_ref):
        # "C5" -> 2 (zero-based column). Letters are base-26 (A=0).
        letters = re.match(r"[A-Z]+", cell_ref).group(0)
        idx = 0
        for ch in letters:
            idx = idx * 26 + (ord(ch) - ord("A") + 1)
        return idx - 1

    grid = []
    for row in sheet.iter("{%s}row" % ns["m"]):
        cells = {}
        for c in row.findall("m:c", ns):
            v = c.find("m:v", ns)
            if v is None:
                text = ""
            elif c.get("t") == "s":          # shared-string cell
                text = shared[int(v.text)]
            elif c.get("t") == "inlineStr":  # rare inline string
                text = "".join(t.text or "" for t in c.iter("{%s}t" % ns["m"]))
            else:                            # number/bool/etc. -> raw text
                text = v.text
            cells[col_index(c.get("r"))] = text
        width = (max(cells) + 1) if cells else 0
        grid.append([cells.get(i, "") for i in range(width)])

    grid = [r for r in grid if any(str(c).strip() for c in r)]
    headers = [str(h).strip() for h in grid[0]]
    return headers, grid[1:]


def read_projects_sheet():
    """Return (headers, rows) for the Projects sheet, preferring openpyxl."""
    try:
        import openpyxl  # noqa: F401
        return read_xlsx_openpyxl(XLSX)
    except Exception as exc:  # ImportError, or any openpyxl read failure
        print(f"[catalog] openpyxl unavailable ({exc}); using stdlib xlsx reader.")
        return read_xlsx_stdlib(XLSX)


def parse_deepdive(path: Path) -> dict:
    """Split the DeepDive markdown into one verbatim block per project ID.

    Project entries look like:  '### 1.1 Molecular Dynamics Engine ... Established'
    and run until the next '### ' header or a '---' rule. We key each block by
    its ID ('1.1') so records can attach their full prose.
    """
    blocks, cur_id, cur = {}, None, []
    header_re = re.compile(r"^###\s+(\d+\.\d+)\b")
    for line in path.read_text(encoding="utf-8").splitlines():
        m = header_re.match(line)
        if m:
            if cur_id:
                blocks[cur_id] = "\n".join(cur).strip()
            cur_id, cur = m.group(1), [line]
        elif cur_id is not None:
            if line.strip() == "---":
                blocks[cur_id] = "\n".join(cur).strip()
                cur_id, cur = None, []
            else:
                cur.append(line)
    if cur_id:
        blocks[cur_id] = "\n".join(cur).strip()
    return blocks


def build_records():
    """Merge xlsx columns + deepdive prose into the list of catalog records."""
    headers, rows = read_projects_sheet()
    # Map header text -> column index so we are robust to column reordering.
    col = {name: i for i, name in enumerate(headers)}
    required = ["Section", "Domain", "ID", "Project", "Difficulty", "Maturity",
                "Deep Dive", "Key Algorithms", "Datasets",
                "Starter Repos / Tools", "CUDA Libraries & GPU Pattern"]
    missing = [c for c in required if c not in col]
    if missing:
        sys.exit(f"[catalog] FATAL: Projects sheet missing columns: {missing}\n"
                 f"          headers seen: {headers}")

    deepdive = parse_deepdive(DEEPDIVE)

    def cell(row, name):
        v = row[col[name]]
        return "" if v is None else str(v).strip()

    records, seen_paths = [], {}
    for row in rows:
        id_str = cell(row, "ID")
        if not id_str:
            continue
        section = int(float(cell(row, "Section")))  # tolerate "1" or "1.0"
        project = cell(row, "Project")
        difficulty = cell(row, "Difficulty")
        slug = slugify(project)
        padded = pad_id(id_str)
        domain_slug = DOMAIN_SLUGS[section]
        folder = f"projects/{domain_slug}/{padded}-{slug}"

        # Guard against the unlikely event of two projects slugging identically
        # inside one domain -- folder names MUST be unique.
        if folder in seen_paths:
            sys.exit(f"[catalog] FATAL: folder collision {folder} "
                     f"(IDs {seen_paths[folder]} and {id_str})")
        seen_paths[folder] = id_str

        records.append({
            "section": section,
            "domain": cell(row, "Domain"),
            "domain_slug": domain_slug,
            "id": id_str,
            "id_padded": padded,
            "project": project,
            "project_slug": slug,
            "difficulty": difficulty,
            "difficulty_emoji": DIFF_EMOJI.get(difficulty, ""),
            "maturity": cell(row, "Maturity"),
            "deep_dive": cell(row, "Deep Dive"),
            "algorithms": cell(row, "Key Algorithms"),
            "datasets": cell(row, "Datasets"),
            "repos": cell(row, "Starter Repos / Tools"),
            "cuda_gpu": cell(row, "CUDA Libraries & GPU Pattern"),
            "deepdive_md": deepdive.get(id_str, ""),
            "folder_path": folder,
        })

    # Stable sort: by section, then numeric minor ID, so catalog.json order
    # matches the on-disk folder order (1.01, 1.02, ... 2.01, ...).
    records.sort(key=lambda r: (r["section"], int(r["id"].split(".")[1])))
    return records


def validate(records):
    """Loud tripwires: counts, section coverage, deepdive attachment."""
    problems = []
    if len(records) != EXPECTED_PROJECTS:
        problems.append(f"expected {EXPECTED_PROJECTS} projects, got {len(records)}")
    sections = {r["section"] for r in records}
    if sections != set(range(1, EXPECTED_SECTIONS + 1)):
        problems.append(f"sections present {sorted(sections)} != 1..{EXPECTED_SECTIONS}")
    no_prose = [r["id"] for r in records if not r["deepdive_md"]]
    if no_prose:
        # Non-fatal but worth surfacing -- THEORY authors lose the prose crutch.
        print(f"[catalog] WARNING: {len(no_prose)} projects have no deepdive block: "
              f"{no_prose[:8]}{' ...' if len(no_prose) > 8 else ''}")
    if problems:
        sys.exit("[catalog] VALIDATION FAILED:\n  - " + "\n  - ".join(problems))


def main():
    ap = argparse.ArgumentParser(description="Build catalog.json from the xlsx + deepdive md.")
    ap.add_argument("--check", action="store_true", help="parse + validate only; write nothing")
    args = ap.parse_args()

    if not XLSX.exists():
        sys.exit(f"[catalog] FATAL: missing {XLSX.name} at repo root.")
    if not DEEPDIVE.exists():
        sys.exit(f"[catalog] FATAL: missing {DEEPDIVE.name} at repo root.")

    records = build_records()
    validate(records)

    by_domain = {}
    for r in records:
        by_domain.setdefault(r["domain"], 0)
        by_domain[r["domain"]] += 1

    if args.check:
        print(f"[catalog] OK (check only): {len(records)} projects, "
              f"{len(by_domain)} domains. No file written.")
        return

    OUT.write_text(json.dumps(records, indent=2, ensure_ascii=False) + "\n",
                   encoding="utf-8")
    print(f"[catalog] Wrote {OUT.relative_to(ROOT)} : {len(records)} projects "
          f"across {len(by_domain)} domains.")
    for dom, n in by_domain.items():
        print(f"           {n:>3}  {dom}")


if __name__ == "__main__":
    main()
